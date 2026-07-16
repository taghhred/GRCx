"""Parse standardized GRCx Risk Assessment Excel workbooks."""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
MAX_UPLOAD_BYTES = 25 * 1024 * 1024
UNSAFE_NAME = re.compile(r'[\\/<>:"|?*]')

HEADER_ALIASES = {
    "risk id": "risk_id",
    "risk title": "title",
    "risk category": "category",
    "affected asset": "affected_asset",
    "business unit": "business_unit",
    "department": "department",
    "vendor": "vendor",
    "risk owner": "owner",
    "owner": "owner",
    "risk scenario / description": "description",
    "risk scenario": "description",
    "description": "description",
    "inherent likelihood": "inherent_likelihood",
    "inherent impact": "inherent_impact",
    "inherent risk score": "inherent_score",
    "inherent risk level": "inherent_level",
    "treatment strategy": "treatment",
    "treatment": "treatment",
    "planned risk controls": "planned_controls",
    "applicable framework": "framework",
    "framework": "framework",
    "framework control reference": "framework_control_ref",
    "residual likelihood": "residual_likelihood",
    "residual impact": "residual_impact",
    "residual risk score": "residual_score",
    "residual risk level": "residual_level",
    "risk workflow status": "status",
    "status": "status",
    "date identified": "date_identified",
    "next review date": "next_review_date",
    "evidence id": "evidence_code",
    "additional notes": "notes",
    "notes": "notes",
}


@dataclass
class ParsedRiskRow:
    risk_id: str
    title: str
    category: str
    affected_asset: str | None
    business_unit: str | None
    department: str | None
    vendor: str | None
    owner: str
    description: str | None
    inherent_likelihood: float | None
    inherent_impact: float | None
    inherent_score: float | None
    inherent_level: str | None
    treatment: str | None
    planned_controls: str | None
    framework: str | None
    framework_control_ref: str | None
    residual_likelihood: float | None
    residual_impact: float | None
    residual_score: float | None
    residual_level: str | None
    status: str
    date_identified: date | None
    next_review_date: date | None
    evidence_code: str | None
    notes: str | None


@dataclass
class ParseResult:
    rows: list[ParsedRiskRow]
    errors: list[str]
    sheet_name: str


def safe_filename(name: str) -> str:
    base = Path(name).name.strip()
    if not base or UNSAFE_NAME.search(base) or ".." in base:
        raise ValueError("Unsafe filename")
    ext = Path(base).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError("Only .xlsx / .xls files are accepted")
    return re.sub(r"\s+", "_", base)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def detect_vendor_from_filename(filename: str) -> str | None:
    lower = filename.lower()
    for vendor in ("ibm", "microsoft", "splunk", "aws", "google", "oracle"):
        if vendor in lower:
            return vendor.title() if vendor != "ibm" else "IBM"
    return None


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(rows[:30]):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            if cell is None:
                continue
            key = HEADER_ALIASES.get(str(cell).strip().lower())
            if key:
                mapping[key] = col
        if "risk_id" in mapping and "title" in mapping:
            return idx, mapping
    return None


def parse_risk_register_bytes(data: bytes) -> ParseResult:
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValueError("File exceeds maximum upload size")
    if len(data) < 64:
        raise ValueError("File is empty or invalid")

    # Reject obvious executables
    if data[:2] == b"MZ" or data[:4] == b"\x7fELF":
        raise ValueError("Executable uploads are not allowed")

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        sheet_name = "Risk Register" if "Risk Register" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    found = _find_header_row(rows)
    if not found:
        return ParseResult(rows=[], errors=["Risk Register header row not found"], sheet_name=sheet_name)

    header_idx, mapping = found
    parsed: list[ParsedRiskRow] = []
    errors: list[str] = []

    for row_num, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        def cell(key: str) -> Any:
            col = mapping.get(key)
            if col is None or col >= len(row):
                return None
            return row[col]

        risk_id = _as_str(cell("risk_id"))
        title = _as_str(cell("title"))
        if not risk_id and not title:
            continue
        if not risk_id:
            errors.append(f"Row {row_num}: missing Risk ID")
            continue
        if not title:
            errors.append(f"Row {row_num}: missing Risk Title for {risk_id}")
            continue

        business_unit = _as_str(cell("business_unit"))
        department = _as_str(cell("department")) or business_unit
        inherent_l = _as_float(cell("inherent_likelihood"))
        inherent_i = _as_float(cell("inherent_impact"))
        inherent_score = _as_float(cell("inherent_score"))
        if inherent_score is None and inherent_l is not None and inherent_i is not None:
            inherent_score = inherent_l * inherent_i
        residual_l = _as_float(cell("residual_likelihood"))
        residual_i = _as_float(cell("residual_impact"))
        residual_score = _as_float(cell("residual_score"))
        if residual_score is None and residual_l is not None and residual_i is not None:
            residual_score = residual_l * residual_i

        parsed.append(
            ParsedRiskRow(
                risk_id=risk_id,
                title=title,
                category=_as_str(cell("category")) or "General",
                affected_asset=_as_str(cell("affected_asset")),
                business_unit=business_unit,
                department=department,
                vendor=_as_str(cell("vendor")),
                owner=_as_str(cell("owner")) or "Unassigned",
                description=_as_str(cell("description")),
                inherent_likelihood=inherent_l,
                inherent_impact=inherent_i,
                inherent_score=inherent_score,
                inherent_level=_as_str(cell("inherent_level")),
                treatment=_as_str(cell("treatment")),
                planned_controls=_as_str(cell("planned_controls")),
                framework=_as_str(cell("framework")),
                framework_control_ref=_as_str(cell("framework_control_ref")),
                residual_likelihood=residual_l,
                residual_impact=residual_i,
                residual_score=residual_score,
                residual_level=_as_str(cell("residual_level")),
                status=_as_str(cell("status")) or "Open",
                date_identified=_as_date(cell("date_identified")),
                next_review_date=_as_date(cell("next_review_date")),
                evidence_code=_as_str(cell("evidence_code")),
                notes=_as_str(cell("notes")),
            )
        )

    return ParseResult(rows=parsed, errors=errors, sheet_name=sheet_name)


def parse_risk_register_path(path: Path) -> ParseResult:
    return parse_risk_register_bytes(path.read_bytes())
